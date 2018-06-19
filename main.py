# AUTOBROWSER (AB)
import splinter
import selenium
import time
import re
# /AB

# EXCEL
from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill
# /EXCEL

# AB
from selenium.common.exceptions import TimeoutException
# from utils import clean_wb

# Set driver variables
webdriver = selenium.webdriver
Browser = splinter.Browser

    # Set chrome options
prefs = {
    'credentials_enable_service': False,
    'profile': {
        'password_manager_enabled': False,
        'default_content_setting_values' : {
            'automatic_downloads': True,
        },

    },
}
chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option("prefs",prefs)
chrome_options.add_argument('--window-size=1200,1000')
# /AB

def auto_browse(fail_tuple,search_list):
    # EXCEL
    wb = load_workbook('Ads_Crawler.xlsx')
    ws = wb['Input']
    ws2 = wb['Output']
    url_list = []
    src = ""
    html = ""

    # clean_wb(wb,ws,ws2)
    # wb = load_workbook('Ads_Crawler.xlsx')
    # ws = wb['Sheet1']
    # search_set = set()
    # url_list = []
        # Construct search_set
    # for cell in ws['F']:
    #    if cell.value and cell.value != u'Search for:':
    #        search_set.add(cell.value)
    # /EXCEL

    # EXCEL
    # for row in ws['A{}:A{}'.format(ws.min_row + 1, ws.max_row)]:
    #     for cell in row:
    #         url_list.append("http://{0}/ads.txt".format( (cell.value).encode('utf-8').strip() ))

    # Add one to list length to accound for start=2 below
    max_row = len(fail_tuple) - 1

    # /EXCEL
    # for row, url in enumerate(url_list, start=2):
    for index, row_url in enumerate(fail_tuple):
        row = row_url[0]
        url = row_url[1]

        print "____________________________"
        print url
        print ""

        if index == 0:
            # Set browser variable
            browser = Browser('chrome', options=chrome_options)

            # Create driver variable for selenium
            # Set timeout limit
            selenium_driver = browser.driver
            selenium_driver.set_page_load_timeout(15)

            # # Open chrome
            # try:
            #     browser.visit(url)
            #
            #     counter = 0
            #     # Test for "pre" element
            #     while browser.is_element_not_present_by_tag("pre",wait_time=1):
            #         counter = counter + 1
            #
            #         if counter > 15:
            #             break
            #         else:
            #             continue
            #
            # except TimeoutException:
            #     pass

        else:
            # Close old tab
            browser.windows[0].close()

            # Switch current window to new tab
            browser.windows.current = browser.windows[0]

        # Visit URL
        try:
            browser.visit(url)
            counter = 0
            # Test for "pre" element
            while browser.is_element_not_present_by_tag("pre",wait_time=1):
                counter = counter + 1

                if counter > 15:
                    break
                else:
                    continue

        except TimeoutException:
            pass

        if browser.is_element_present_by_tag("pre"):

            # Remove manual check flag from Notes column
            ws['{0}{1}'.format("D",row)] = ""
            src = selenium_driver.page_source

            # Loop through search_set from wb
            for col, search_str in enumerate(search_list, start=2):

                # Check for desired text
                text_found = re.search(r'(\<|\,| ){0}(\>|\,| )'.format(search_str), src)

                # Confirm if found and choose column accordingly
                if text_found:
                    # Input tab
                    column = "B"

                    # Output tab
                    response = "Pass"
                    colorFill = PatternFill(start_color='00FF00',
                       end_color='00FF00',
                       fill_type='solid')

                    # Log response to terminal
                    print "{0} found!".format(search_str)
                else:
                    # Input tab
                    column = "C"

                    # Output tab
                    response = "Fail"
                    colorFill = PatternFill(start_color='FF0000',
                       end_color='FF0000',
                       fill_type='solid')

                    # Log response to terminal
                    print "{0} not found!".format(search_str)

                # Populate 'Output' tab body
                ws2.cell(row=row,column=col).value = response
                ws2.cell(row=row,column=col).fill= colorFill

                # Populate spreadsheet using column above
                if ws['{0}{1}'.format(column,row)].value:
                    current_val = ws['{0}{1}'.format(column,row)].value
                    ws['{0}{1}'.format(column,row)] = "{0}, {1}".format(current_val, search_str)
                else:
                    ws['{0}{1}'.format(column,row)] = search_str

        print "____________________________"

        # Write confirmation to file
        wb.save("Ads_Crawler.xlsx")

        # Open new tab
        selenium_driver.execute_script("window.open('');")

        if index == max_row:
            # Close browser
            browser.quit()
