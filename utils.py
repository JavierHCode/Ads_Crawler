from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter

def clean_wb(wb,ws,ws2):
    clearFill = PatternFill(fill_type=None)

    for row in ws['B2:D{0}'.format(ws.max_row)]:
      for cell in row:
        cell.value = None

    for row in ws2['A1:{0}{1}'.format(get_column_letter(ws2.max_column),ws2.max_row)]:
      for cell in row:
        cell.value = None
        cell.fill = clearFill

    wb.save("Ads_Crawler.xlsx")
