# -*- coding: utf-8 -*-

# Import dependencies
import cchardet as chardet
import cookielib
import httplib2
import httplib
import urllib2
import socket
import time
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.fills import PatternFill
from bs4 import BeautifulSoup, UnicodeDammit
from utils import clean_wb
from main import auto_browse
import ssl

# Configure opener to accept cookies and handle HTTP/HTTPS
cookies = cookielib.LWPCookieJar()
handlers = [
    urllib2.HTTPHandler(),
    urllib2.HTTPSHandler(),
    urllib2.HTTPCookieProcessor(cookies)
    ]
opener = urllib2.build_opener(*handlers)
opener.addheaders = [('User-Agent', 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.107 Safari/537.36'),('Accept', '*/*')]
urllib2.install_opener(opener)

# # Compile RegEx for string matching
# re.compile

# Define helper functions
def meta_redirect(content):
    soup  = BeautifulSoup(content, 'html.parser' )
    result=soup.find("meta",attrs={"http-equiv":"refresh"})
    if result:
        wait,text=result["content"].split(";")
        if text.strip().lower().startswith("url="):
            url=text[4:]
            return url
    return None

def get_content(url):

    h=httplib2.Http(".cache")
    # h.add_certificate(key,cert,"")

    resp, content = h.request(url,"GET")

    # follow the chain of redirects
    while meta_redirect(content):
        resp, content = h.request(meta_redirect(content),"GET")

    return content

# def clean_wb(wb,ws,ws2):
#
#     clearFill = PatternFill(fill_type=None)
#
#     for row in ws['B2:D{0}'.format(ws.max_row)]:
#       for cell in row:
#         cell.value = None
#
#     for row in ws2['A1:{0}{1}'.format(get_column_letter(ws2.max_column),ws2.max_row)]:
#       for cell in row:
#         cell.value = None
#         cell.fill = clearFill
#
#     wb.save("Ads_Crawler.xlsx")

# EXCEL
wb = load_workbook('Ads_Crawler.xlsx')
ws = wb['Input']
ws2 = wb['Output']
search_list = []
url_list = []
src = ""
html = ""

clean_wb(wb,ws,ws2)

# Construct search_list
for col, cell in enumerate(ws['F'], start=1):
   if cell.value and cell.value != u'Search for:':
       cell_str = (str(cell.value)).encode('utf-8').strip()
       search_list.append(cell_str)

       # Populate 'Output' tab headers
       ws2.cell(row=1,column=col).value = cell_str

# /EXCEL

# EXCEL
# Construct url_list
for row in ws['A{}:A{}'.format(ws.min_row + 1, ws.max_row)]:
    for cell in row:
        if cell.value:
            url_list.append("http://{0}/ads.txt".format( (cell.value).encode('utf-8').strip()) )

# Add one to list length to accound for start=2 below
max_row = len(url_list) + 1
fail_tuple = []
# /EXCEL

# Loop through url_list with index starting at 2 to match row index in spreadsheet
for row, url in enumerate(url_list, start=2):
    print "____________________________"
    print url
    print ""

    ws2['{0}{1}'.format("A",row)] = url

    # Attempt to open URL
    try:
        html = urllib2.urlopen(url, timeout = 60)#"http://cronista.com/ads.txt"/"http://mundodrama.com/ads.txt"
        html = html.read()

        print chardet.detect(html)

        # Check for refresh chain and handle accordingly w/ urllib2 or httplib2
        if "refresh" in html.lower():
            src = get_content(url)
        else:
            soup = BeautifulSoup( html, 'html.parser', from_encoding="windows-1255")
            src = soup.get_text()

    # Except for errors and print to terminal
    except httplib.BadStatusLine as e:
        print e.reason #http error
        src = ""
        html = ""
    except urllib2.URLError as e:
        print e.reason    #not catch
        src = ""
        html = ""
    except socket.timeout as e:
        print e    #catched
        src = ""
        html = ""
    except ssl.SSLError as e:
        print e    #catched
        src = ""
        html = ""

    # Log failed URLs in Notes section
    if chardet.detect(html)['encoding'] == None and chardet.detect(html)['confidence'] == None:
        print "Confidence and Encoding for {0} are None".format(url)
        ws['{0}{1}'.format("D",row)] = "Verify URL manually."

        # Add failed row and url to fail_tuple for auto_browsing
        fail_url = (row,url)
        fail_tuple.append( fail_url )

    # If successful, continue with searches
    else:
        # Loop through search_set from wb
        for col, search_str in enumerate(search_list, start=2):
        # for search_str in search_set:
            # Check for desired text
            text_found = re.search(r'(\<|\,| ){0}(\>|\,| )'.format(search_str), src)

            # Confirm if found and choose column accordingly
            if text_found:
                # Input tab
                column = "B"

                # Output tab
                response = "Pass"
                colorFill = PatternFill(start_color='00FF00',
                   end_color='00FF00',
                   fill_type='solid')

                # Log response to terminal
                print "{0} found!".format(search_str)
            else:
                # Input tab
                column = "C"

                # Output tab
                response = "Fail"
                colorFill = PatternFill(start_color='FF0000',
                   end_color='FF0000',
                   fill_type='solid')

                # Log response to terminal
                print "{0} not found!".format(search_str)

            # Populate 'Output' tab body
            ws2.cell(row=row,column=col).value = response
            ws2.cell(row=row,column=col).fill= colorFill

            # Populate spreadsheet using column above
            if ws['{0}{1}'.format(column,row)].value:
                current_val = ws['{0}{1}'.format(column,row)].value
                ws['{0}{1}'.format(column,row)] = "{0}, {1}".format(current_val, search_str)
            else:
                ws['{0}{1}'.format(column,row)] = search_str

    print "____________________________"

    # Write confirmation to file
    wb.save("Ads_Crawler.xlsx")
    # Run autobrowser to reduce fail list

print "Fail tuple: ", fail_tuple
print "Search list: ", search_list

# fail_tuple = [(45, 'http://cronista.com/ads.txt'),(117, 'http://eluniversal.com/ads.txt'), (122, 'http://entrepreneur.com/es/ads.txt'), (123, 'http://esquirelat.com/ads.txt'), (124, 'http://estampas.com/ads.txt'),
# (135, 'http://FutbolAlReves.com/ads.txt'), (144, 'http://goal.com/es/ads.txt'), (149, 'http://holaciudad.com/ads.txt'), (151, 'http://hoylosangeles.com/ads.txt'), (165, 'http://lalupa.com.ve/ads.txt'), (169, 'http://lapaginamillionaria.com/ads.txt'), (170, 'http://LaPaginaMillonaria.com\xc2\xa0/ads.txt'), (177, 'http://larevista.ec/ads.txt'), (179, 'http://LaSeleccion.com.ar/ads.txt'), (192, 'http://lostiempos.com/ads.txt'), (208, 'http://minuto30.com/ads.txt'),
# (217, 'http://mundodrama.com/ads.txt'), (229, 'http://nuevamujer.com/ads.txt'), (234, 'http://orlandosentinel.com/elsentinel/ads.txt'), (252, 'http://quevidavideo.com/ads.txt'), (256, 'http://rankingsdefutbol.com/ads.txt'), (267, 'http://runrun.es/ads.txt'), (272, 'http://sinembargo.mx/ads.txt'), (285, 'http://tuvisioncanal.com/ads.txt'), (286, 'http://tvmax-9.com/ads.txt'), (287, 'http://tvn-2.com/ads.txt'), (295, 'http://VamosPeruanos.com/ads.txt'),
# (301, 'http://vivelohoy.com/ads.txt'), (306, 'http://yoamoloszapatos.com/ads.txt')]

# search_list = ['pub-4769180667814197', 'pub-7995104076770938', '120', '184081', '13894', '4aa37fe0', '282', '282', '185362', '95838', '95838', 'ey5io-qyxzx', '148395', '148395', '232757', '232757', '538959099', '137711', '156212',
# '20535', '101557', '15777', '5322', '16156', '8035', '156500', '8790', '18008', '247572', '247572', '1360', '11645', '539924617', '156700', '17960', 'pub-1956856209985681', '82069', '82069', '762737', '762769', '493', '527', '1356', '1fe16a48',
# '10440', '20535']

auto_browse(fail_tuple,search_list)
